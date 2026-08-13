"""Parses and validates the two Excel templates (Warehouse Master, Location
Master) and creates/updates Warehouse and Location rows through the same
id_generator-backed services used by the manual API (warehouse_service,
location_service) -- so every code an upload produces is generated exactly
the same way a manually-created one would be, never a separate code path.

Partial-success by design: a file with a few bad rows still commits
everything else and returns a precise per-row error report, matching the
pattern from the sibling WMS Readiness Tracker project.

Column choice deliberately does NOT match the original brief's illustrative
"Zone/Rack/Bay/Level/Bin" Location Master example -- the REAL formula,
confirmed against real company data (see id_generator_service.py), uses
Category Rack + a generated sequence instead. Real files took priority over
the brief's placeholder text once they were provided.
"""
import io
import uuid
from dataclasses import dataclass

import openpyxl
import pandas as pd
from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.category_rack_mapping import CategoryRackMapping
from app.models.location import Location
from app.models.location_type import LocationTypeConfig
from app.models.site import Site
from app.models.upload_batch import UploadBatch
from app.models.upload_error import UploadError
from app.models.warehouse import Warehouse
from app.services.location_service import create_location
from app.services.merge_suggestion_service import create_suggestion, find_similar_location
from app.services.warehouse_service import create_warehouse

MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_ROWS = 20_000
MAX_COLUMNS = 50


class UploadRejected(Exception):
    """Whole file rejected before any row was processed -- bad type/size, or a
    required column is missing entirely. Distinct from a row-level error."""


@dataclass
class RowError:
    row_number: int
    column_name: str
    error_message: str


def _normalize_header(value) -> str:
    return " ".join(str(value).strip().lower().split())


def _check_file_basics(filename: str, content: bytes) -> None:
    if not filename.lower().endswith(".xlsx"):
        raise UploadRejected("Only .xlsx files are accepted")
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise UploadRejected(f"File is too large (max {MAX_FILE_SIZE_BYTES // (1024 * 1024)}MB)")


def _check_workbook_dimensions(content: bytes) -> None:
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    try:
        sheet = workbook.active
        if sheet.max_row and sheet.max_row > MAX_ROWS:
            raise UploadRejected(f"Sheet has too many rows (max {MAX_ROWS})")
        if sheet.max_column and sheet.max_column > MAX_COLUMNS:
            raise UploadRejected(f"Sheet has too many columns (max {MAX_COLUMNS})")
    finally:
        workbook.close()


def _read_dataframe(content: bytes) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(content), engine="openpyxl", dtype=str)


def _require_columns(df: pd.DataFrame, required: dict[str, str]) -> dict[str, str]:
    available = {_normalize_header(col): col for col in df.columns}
    missing = [display for norm, display in required.items() if norm not in available]
    if missing:
        raise UploadRejected(f"Missing required column(s): {', '.join(missing)}")
    return {norm: available[norm] for norm in required}


def _cell(row: pd.Series, column: str) -> str | None:
    value = row.get(column)
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    return text or None


def _new_batch(db: Session, uploaded_by: uuid.UUID, filename: str, file_type: str) -> UploadBatch:
    batch = UploadBatch(
        file_type=file_type, uploaded_by=uploaded_by, file_name=filename,
        row_count=0, success_count=0, error_count=0, status="processing",
    )
    db.add(batch)
    db.flush()
    return batch


def _reject(db: Session, batch: UploadBatch, reason: str) -> UploadBatch:
    batch.status = "failed"
    batch.error_count = 1
    db.add(UploadError(batch_id=batch.id, row_number=0, column_name="(file)", error_message=reason))
    db.commit()
    db.refresh(batch)
    return batch


def _finalize(
    db: Session, batch: UploadBatch, row_count: int, success_count: int, errors: list[RowError], pending_count: int = 0
) -> UploadBatch:
    batch.row_count = row_count
    batch.success_count = success_count
    batch.error_count = len(errors)
    batch.pending_count = pending_count
    batch.status = "completed"
    for err in errors:
        db.add(UploadError(batch_id=batch.id, row_number=err.row_number, column_name=err.column_name, error_message=err.error_message))
    db.commit()
    db.refresh(batch)
    return batch


# --- Warehouse Master ---


def ingest_warehouse_master(db: Session, uploaded_by: uuid.UUID, filename: str, content: bytes) -> UploadBatch:
    """Upsert by (site, warehouse_type_code, warehouse_code) -- the formula's
    own natural key. An existing warehouse's name/description/capacity are
    updated in place on a re-upload; its generated_code/duplicate_letter are
    left untouched (those are only ever decided by create_warehouse's
    relettering logic, on genuinely new warehouses)."""
    batch = _new_batch(db, uploaded_by, filename, "warehouse_master")

    try:
        _check_file_basics(filename, content)
        _check_workbook_dimensions(content)
        df = _read_dataframe(content)
        columns = _require_columns(
            df,
            {
                "site code": "Site Code",
                "warehouse type code": "Warehouse Type Code",
                "warehouse code": "Warehouse Code",
                "warehouse name": "Warehouse Name",
            },
        )
    except UploadRejected as exc:
        return _reject(db, batch, str(exc))

    optional = {_normalize_header(c): c for c in df.columns}
    description_col = optional.get("description")
    capacity_col = optional.get("capacity")

    errors: list[RowError] = []
    success_count = 0
    site_cache: dict[str, Site | None] = {}

    for i, row in df.iterrows():
        row_number = i + 2
        site_code = _cell(row, columns["site code"])
        wh_type = _cell(row, columns["warehouse type code"])
        wh_code = _cell(row, columns["warehouse code"])
        name = _cell(row, columns["warehouse name"])
        description = _cell(row, description_col) if description_col else None
        capacity_raw = _cell(row, capacity_col) if capacity_col else None

        row_ok = True
        if not name:
            errors.append(RowError(row_number, "Warehouse Name", "Warehouse Name is required"))
            row_ok = False
        if not site_code:
            errors.append(RowError(row_number, "Site Code", "Site Code is required"))
            row_ok = False
        if not wh_type:
            errors.append(RowError(row_number, "Warehouse Type Code", "Warehouse Type Code is required"))
            row_ok = False
        if not wh_code:
            errors.append(RowError(row_number, "Warehouse Code", "Warehouse Code is required"))
            row_ok = False

        capacity: int | None = None
        if capacity_raw:
            try:
                capacity = int(capacity_raw)
            except ValueError:
                errors.append(RowError(row_number, "Capacity", f"'{capacity_raw}' is not a whole number"))
                row_ok = False

        if not row_ok:
            continue

        if site_code not in site_cache:
            site_cache[site_code] = db.scalar(select(Site).where(Site.code == site_code))
        site = site_cache[site_code]
        if site is None:
            errors.append(RowError(row_number, "Site Code", f"No site with code '{site_code}'"))
            continue

        # Upsert key includes `name`, not just (site, type, code) -- the
        # formula explicitly allows MULTIPLE warehouses to share one
        # (type, code) pair, disambiguated only by the auto-assigned
        # duplicate_letter (e.g. "Pharmacy Mainstore Drugs" vs "...
        # Consumables" both under type A code 01). Name is the only column
        # available to tell those apart on a re-upload; matching on
        # (type, code) alone would silently collapse two real warehouses
        # into one on a second upload.
        existing = db.scalar(
            select(Warehouse).where(
                Warehouse.site_id == site.id,
                Warehouse.warehouse_type_code == wh_type,
                Warehouse.warehouse_code == wh_code,
                Warehouse.name == name,
            )
        )
        if existing is not None:
            existing.name = name
            existing.description = description
            existing.capacity = capacity
            success_count += 1
            continue

        try:
            create_warehouse(db, site.id, wh_type, wh_code, name, description, capacity)
            success_count += 1
        except HTTPException as exc:
            errors.append(RowError(row_number, "Warehouse Type Code / Warehouse Code", exc.detail))

    return _finalize(db, batch, len(df), success_count, errors)


# --- Location Master ---


def ingest_location_master(db: Session, uploaded_by: uuid.UUID, filename: str, content: bytes) -> UploadBatch:
    """Matches each row to an existing Warehouse by its already-generated
    `generated_code` (the natural, human-visible identifier once warehouses
    exist). Two layers of duplicate handling, deliberately different:
    an EXACT (warehouse, location_type, description) repeat is skipped
    outright as an error -- nothing to do, it's already there. A SIMILAR
    but not exact description (e.g. differently-worded but likely the same
    real place) is held as a pending MergeSuggestion instead of being
    auto-created OR silently rejected -- see merge_suggestion_service."""
    batch = _new_batch(db, uploaded_by, filename, "location_master")

    try:
        _check_file_basics(filename, content)
        _check_workbook_dimensions(content)
        df = _read_dataframe(content)
        columns = _require_columns(
            df,
            {
                "warehouse code": "Warehouse Code",
                "category rack": "Category Rack",
                "description": "Description",
            },
        )
    except UploadRejected as exc:
        return _reject(db, batch, str(exc))

    errors: list[RowError] = []
    success_count = 0
    pending_count = 0
    warehouse_cache: dict[str, Warehouse | None] = {}

    for i, row in df.iterrows():
        row_number = i + 2
        warehouse_code = _cell(row, columns["warehouse code"])
        category_rack = _cell(row, columns["category rack"])
        description = _cell(row, columns["description"])

        row_ok = True
        if not warehouse_code:
            errors.append(RowError(row_number, "Warehouse Code", "Warehouse Code is required"))
            row_ok = False
        if not category_rack:
            errors.append(RowError(row_number, "Category Rack", "Category Rack is required"))
            row_ok = False
        if not description:
            errors.append(RowError(row_number, "Description", "Description is required"))
            row_ok = False
        if not row_ok:
            continue

        if warehouse_code not in warehouse_cache:
            warehouse_cache[warehouse_code] = db.scalar(select(Warehouse).where(Warehouse.generated_code == warehouse_code))
        warehouse = warehouse_cache[warehouse_code]
        if warehouse is None:
            errors.append(RowError(row_number, "Warehouse Code", f"No warehouse with code '{warehouse_code}'"))
            continue

        mapping = db.scalar(
            select(CategoryRackMapping).where(
                CategoryRackMapping.warehouse_type_code == warehouse.warehouse_type_code,
                CategoryRackMapping.raw_category_text == category_rack.upper(),
            )
        )
        if mapping is None:
            errors.append(
                RowError(
                    row_number, "Category Rack",
                    f"'{category_rack}' is not a configured Category Rack mapping for warehouse type "
                    f"'{warehouse.warehouse_type_code}'",
                )
            )
            continue
        loc_type = db.get(LocationTypeConfig, mapping.location_type_config_id)

        already_exists = db.scalar(
            select(Location).where(
                Location.warehouse_id == warehouse.id,
                Location.location_type_code == loc_type.code,
                Location.description == description,
            )
        )
        if already_exists is not None:
            errors.append(
                RowError(row_number, "Description", f"Duplicate: '{description}' already exists for this warehouse")
            )
            continue

        similar_location, similarity = find_similar_location(db, warehouse.id, loc_type.code, description)
        if similar_location is not None:
            create_suggestion(
                db, batch.id, row_number, warehouse.id, loc_type.code, category_rack, description,
                similar_location, similarity,
            )
            pending_count += 1
            continue

        try:
            create_location(db, warehouse.id, loc_type.code, category_rack, description)
            success_count += 1
        except HTTPException as exc:
            errors.append(RowError(row_number, "Description", exc.detail))

    return _finalize(db, batch, len(df), success_count, errors, pending_count)
