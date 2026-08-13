"""AI-assisted mapping from raw/legacy hospital data (e.g. RSUS Raw.xlsx --
Organization/CodeStore/Store/CodeStoreRack/StoreRack/ActiveStoreRack, none
of which are the formula's own fields) to the app's real coding scheme.

Same "suggest, never auto-apply" pattern as merge_suggestion_service and
the Revision workflow: nothing here ever creates a Warehouse or Location on
its own. Every suggestion sits `pending` until an admin approves (optionally
overriding the suggested value first) or rejects it -- and the AI is only
ever allowed to choose from ALREADY-CONFIGURED Warehouse Type/Code and
Category Rack values, validated after the fact, never permitted to invent a
new one. That's a hard rule, not just a prompt instruction: any value the
model returns that isn't in the real config gets nulled back out before it's
ever shown to an admin.

THREE rounds of Groq calls happen per raw-import batch, all at upload time
or once per explicit trigger -- never a call per row:
1. Clustering, at upload time -- groups legacy store names that are really
   the SAME physical warehouse under different billing/status labels (e.g.
   "EMERGENCY" + "EMERGENCY NONCHARGEABLE") into one candidate, confirmed
   necessary against real RSUS Mapping.xlsx ground truth: 61 legacy store
   names there collapse into only 48 real final warehouses. Simple keyword
   rules aren't reliable for this (floor variants like "OPD 1ST FLOOR"/
   "OPD 2ND FLOOR" must NOT merge, but billing variants almost always
   should), so it's an AI judgment call, biased toward NOT merging when
   unsure -- see `_suggest_clusters`.
2. Warehouse Type/Code suggestion, at upload time, over the CLUSTERED
   candidates from step 1.
3. Location/Category Rack suggestion, triggered explicitly once warehouse
   review is done (can't run earlier: the valid Category Rack options are
   scoped by the warehouse's real type, which isn't known until its
   suggestion is approved).

Each round is chunked into a small, bounded number of calls
(`_SUGGESTION_CHUNK_SIZE`/`_CLUSTER_CHUNK_SIZE` candidates per call) rather
than one single giant prompt -- confirmed against real data that a single
call covering 60+ candidates in one shot silently drops a handful of them,
which is worse than a few extra calls. Each chunk's response is matched
back to its candidates by an explicit `index` field rather than trusted
response order/count, so even a chunk with a partial response only leaves
THOSE few rows unaddressed, not the whole batch.

Every approved suggestion (whether the AI's own guess or an admin's
override) is fed back into FUTURE suggestion prompts as a confirmed
few-shot example -- "PHARMACY MAINSTORE DRUGS -> A/01, confirmed by an
admin" is a much stronger signal than the model's own prior. This is
retrieval from real data already sitting in the DB, not model
fine-tuning (Groq's hosted API doesn't offer that, and it would be
overkill here) -- see `confirmed_warehouse_examples`/
`confirmed_location_examples`. Examples are drawn from every prior batch
and every site, not just the current one, since the underlying naming
conventions and Warehouse Type/Code meanings are the same across every
Siloam hospital.
"""

import io
import json
import logging
import uuid
from datetime import datetime, timezone

import openpyxl
import pandas as pd
from fastapi import HTTPException, status
from groq import Groq
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.category_rack_mapping import CategoryRackMapping
from app.models.location import Location
from app.models.location_type import LocationTypeConfig
from app.models.raw_import import RawImportBatch, RawLocationSuggestion, RawWarehouseSuggestion
from app.models.site import Site
from app.models.warehouse import Warehouse
from app.models.warehouse_code import WarehouseCodeConfig
from app.services.location_service import create_location
from app.services.merge_suggestion_service import create_suggestion, find_similar_location
from app.services.warehouse_service import create_warehouse

logger = logging.getLogger(__name__)

MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
MAX_ROWS = 20_000
GROQ_MODEL = "llama-3.3-70b-versatile"
# Per-call candidate count -- keeps each prompt/response small enough that
# the model reliably addresses every entry (see module docstring).
_SUGGESTION_CHUNK_SIZE = 40
# How many confirmed examples to feed back into future prompts -- capped so
# the few-shot section stays small relative to the actual candidates being
# suggested; most recent confirmations first, since naming conventions a
# hospital used most recently are the most likely to recur.
_CONFIRMED_EXAMPLE_LIMIT = 60


def _chunks(items: list, size: int) -> list[list]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def confirmed_warehouse_examples(db: Session, limit: int = _CONFIRMED_EXAMPLE_LIMIT) -> list[tuple[str, str, str]]:
    """(legacy_name, warehouse_type_code, warehouse_code) for every
    previously-APPROVED warehouse suggestion, across every batch and site --
    `suggested_warehouse_type_code`/`_code` are overwritten with the final
    applied value on approval (see approve_warehouse_suggestion), whether
    that was the AI's own guess or an admin's override, so this is always
    the real confirmed answer, not just what the model originally
    proposed."""
    rows = list(
        db.scalars(
            select(RawWarehouseSuggestion)
            .where(
                RawWarehouseSuggestion.status == "approved",
                RawWarehouseSuggestion.suggested_warehouse_type_code.isnot(None),
                RawWarehouseSuggestion.suggested_warehouse_code.isnot(None),
            )
            .order_by(RawWarehouseSuggestion.resolved_at.desc())
            .limit(limit)
        ).all()
    )
    return [(r.legacy_name, r.suggested_warehouse_type_code, r.suggested_warehouse_code) for r in rows]


def confirmed_location_examples(db: Session, limit: int = _CONFIRMED_EXAMPLE_LIMIT) -> list[tuple[str, str, str]]:
    """(warehouse_type_code, legacy_description, category_rack) for every
    previously-APPROVED location suggestion, across every batch and site --
    same "final applied value, not just the model's first guess" reasoning
    as confirmed_warehouse_examples."""
    rows = list(
        db.scalars(
            select(RawLocationSuggestion)
            .where(
                RawLocationSuggestion.status == "approved", RawLocationSuggestion.suggested_category_rack.isnot(None)
            )
            .order_by(RawLocationSuggestion.resolved_at.desc())
            .limit(limit)
        ).all()
    )
    result = []
    for r in rows:
        warehouse = db.get(Warehouse, r.warehouse_id)
        if warehouse is not None:
            result.append((warehouse.warehouse_type_code, r.legacy_description, r.suggested_category_rack))
    return result


class UploadRejected(Exception):
    """Whole file rejected before any row was processed."""


# --- parsing ---


def _normalize_header(value) -> str:
    return "".join(str(value).strip().lower().split())


def _cell(row: pd.Series, column: str | None) -> str | None:
    if column is None:
        return None
    value = row.get(column)
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    return text or None


def _parse_raw_rows(content: bytes) -> list[dict]:
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise UploadRejected(f"File is too large (max {MAX_FILE_SIZE_BYTES // (1024 * 1024)}MB)")

    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    try:
        sheet = workbook.active
        if sheet.max_row and sheet.max_row > MAX_ROWS:
            raise UploadRejected(f"Sheet has too many rows (max {MAX_ROWS})")
    finally:
        workbook.close()

    df = pd.read_excel(io.BytesIO(content), engine="openpyxl", dtype=str)
    available = {_normalize_header(c): c for c in df.columns}
    required = {"codestore": "CodeStore", "store": "Store", "codestorerack": "CodeStoreRack", "storerack": "StoreRack"}
    missing = [display for norm, display in required.items() if norm not in available]
    if missing:
        raise UploadRejected(f"Missing required column(s): {', '.join(missing)}")
    active_col = available.get("activestorerack")

    rows = []
    for _, row in df.iterrows():
        code_store = _cell(row, available["codestore"])
        store = _cell(row, available["store"])
        description = _cell(row, available["storerack"])
        if not code_store or not store or not description:
            # Best-effort raw-import aid, not a strictly validated master
            # upload -- an incomplete legacy row is silently dropped rather
            # than blocking the whole file.
            continue
        code_rack = _cell(row, available["codestorerack"])
        active_raw = _cell(row, active_col)
        is_active = active_raw not in ("0", "false", "no")
        rows.append(
            {"code_store": code_store, "store": store, "code_rack": code_rack, "description": description, "is_active": is_active}
        )
    return rows


# --- warehouse-level clustering (consolidation) ---

_CLUSTER_SYSTEM_PROMPT = """You group legacy hospital store names that \
describe the SAME real physical warehouse/department, just labeled \
differently for billing or administrative reasons -- e.g. "X" and \
"X NONCHARGEABLE" (or "X BPJS", "X QUARANTINE", "X REJECTED EXPIRED") are \
almost always the SAME physical place under a different billing/status \
label, not different departments, and should be grouped together.

Do NOT group names that describe genuinely different physical locations, \
even if the words look similar -- e.g. "OPD 1ST FLOOR" and "OPD 2ND FLOOR" \
are different floors and must stay separate; "PHARMACY SATELLITE 2ND \
FLOOR" and "PHARMACY SATELLITE 3RD FLOOR" must stay separate too. When \
genuinely unsure, do NOT group -- leaving a name in its own group of one \
is always safe; a wrong merge is not.

You are given a numbered list of legacy store names. Respond with ONLY a \
JSON object of this exact shape -- every number from 1 to N must appear in \
EXACTLY ONE group:
{"groups": [[1, 3, 8], [2], [4, 5]]}"""
# Names only (no valid-pairs list), so a call comfortably covers far more
# candidates per chunk than the type/code round.
_CLUSTER_CHUNK_SIZE = 120


def _suggest_clusters(candidates: list[dict]) -> list[list[int]]:
    """Returns 0-based index groups covering every candidate exactly once.
    Falls back to every candidate as its own group of one -- i.e. no
    consolidation, identical to this feature's behavior before clustering
    existed -- if no Groq key is configured, the call fails, or a
    candidate is simply never mentioned in the response. A missed or wrong
    NON-merge only costs a slightly less consolidated suggestion queue; a
    wrong merge would silently combine two different real warehouses, so
    every path here is biased toward "leave it separate" as the safe
    default."""
    n = len(candidates)
    if n <= 1 or not settings.groq_api_key:
        return [[i] for i in range(n)]

    client = Groq(api_key=settings.groq_api_key)
    groups: list[list[int]] = []
    seen: set[int] = set()

    for chunk_start, chunk in zip(range(0, n, _CLUSTER_CHUNK_SIZE), _chunks(candidates, _CLUSTER_CHUNK_SIZE)):
        names_text = "\n".join(f"{i + 1}. {c['legacy_name']}" for i, c in enumerate(chunk))
        try:
            response = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {"role": "system", "content": _CLUSTER_SYSTEM_PROMPT},
                    {"role": "user", "content": f"Store names:\n{names_text}"},
                ],
                max_completion_tokens=4000,
                temperature=0.2,
                response_format={"type": "json_object"},
            )
            parsed = json.loads(response.choices[0].message.content or "{}")
            raw_groups = parsed.get("groups")
            if not isinstance(raw_groups, list):
                raise ValueError("Model response didn't match the expected shape")

            for group in raw_groups:
                if not isinstance(group, list) or not group:
                    continue
                global_indices = []
                for local_idx in group:
                    if not isinstance(local_idx, int) or not (1 <= local_idx <= len(chunk)):
                        continue
                    global_idx = chunk_start + local_idx - 1
                    if global_idx in seen:
                        continue  # already placed by an earlier group -- keep the first
                    seen.add(global_idx)
                    global_indices.append(global_idx)
                if global_indices:
                    groups.append(global_indices)
        except Exception:
            logger.exception(
                "Raw import clustering request failed for one chunk -- those stores are left ungrouped (no consolidation)"
            )
            continue

    for i in range(n):
        if i not in seen:
            groups.append([i])

    return groups


def _build_clustered_candidates(candidates: list[dict], groups: list[list[int]]) -> list[dict]:
    """Pure merge step, split out from _suggest_clusters so it's testable
    without a Groq call: each group's members are folded into one
    candidate -- the union of their rack rows, the first member as the
    primary legacy_code/legacy_name, and the rest recorded in
    consolidated_legacy_names for display and for the type/code prompt."""
    merged = []
    for group in groups:
        members = [candidates[i] for i in group]
        primary = members[0]
        merged.append(
            {
                "legacy_code": primary["legacy_code"],
                "legacy_name": primary["legacy_name"],
                "consolidated_legacy_names": [m["legacy_name"] for m in members[1:]],
                "raw_rows": [row for m in members for row in m["raw_rows"]],
            }
        )
    return merged


# --- warehouse-level suggestions ---

_WAREHOUSE_SYSTEM_PROMPT = """You map legacy hospital store names to a fixed \
warehouse coding scheme. You are given a numbered list of legacy store \
names, and the full list of valid (Warehouse Type Code, Warehouse Code) \
pairs with their descriptions. For each store, pick the single best-matching \
valid pair -- you may ONLY use pairs from the list given, never invent one. \
If nothing genuinely fits, use null for both fields. Keep reasoning to a \
short phrase, under 10 words. You MUST return exactly one entry for EVERY \
numbered store, in any order, each carrying its own "index" (the store's \
number) so entries can be matched back correctly even if reordered.

You may also be given a list of confirmed examples -- real mappings an \
admin has already reviewed and approved. Treat an exact or near-exact name \
match against a confirmed example as very strong evidence for the same \
pair, stronger than your own prior guess -- but you must still only pick \
from the valid pairs list.

Respond with ONLY a JSON object of this exact shape:
{"suggestions": [{"index": 1, "warehouse_type_code": "A" or null, "warehouse_code": "01" or null, "reasoning": "..."}]}"""


def _suggest_warehouse_codes(db: Session, candidates: list[dict]) -> list[tuple[str | None, str | None, str]]:
    """Returns one (type_code, code, reasoning) tuple per candidate, same
    order as given. Chunked into `_SUGGESTION_CHUNK_SIZE`-sized calls, and
    within each chunk, matches the model's response back to candidates by
    an explicit `index` field rather than trusting response order/count --
    real models sometimes skip or merge near-duplicate-looking entries
    (e.g. "EMERGENCY" vs "EMERGENCY NONCHARGEABLE"), so a strict
    all-or-nothing length check would silently blank out a whole chunk over
    a handful of misses. Any candidate never addressed just keeps its
    "no suggestion, assign manually" default -- if a Groq key isn't
    configured at all, that's every candidate, and no calls are made."""
    default = (None, None, "AI unavailable -- assign manually")
    result: list[tuple[str | None, str | None, str]] = [default for _ in candidates]
    if not candidates or not settings.groq_api_key:
        return result

    valid_pairs = list(db.scalars(select(WarehouseCodeConfig)).all())
    valid_set = {(p.warehouse_type_code, p.code) for p in valid_pairs}
    options_text = "\n".join(f"- {p.warehouse_type_code}/{p.code}: {p.description}" for p in valid_pairs)

    examples = confirmed_warehouse_examples(db)
    examples_text = ""
    if examples:
        examples_text = "\n\nConfirmed examples (already reviewed and approved by an admin):\n" + "\n".join(
            f"- {name} -> {t}/{c}" for name, t, c in examples
        )

    client = Groq(api_key=settings.groq_api_key)

    for chunk_start, chunk in zip(range(0, len(candidates), _SUGGESTION_CHUNK_SIZE), _chunks(candidates, _SUGGESTION_CHUNK_SIZE)):
        stores_text = "\n".join(
            f"{i + 1}. {c['legacy_name']}"
            + (f" (also covers: {', '.join(c['consolidated_legacy_names'])})" if c.get("consolidated_legacy_names") else "")
            + f" (legacy code {c['legacy_code']})"
            for i, c in enumerate(chunk)
        )
        try:
            response = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {"role": "system", "content": _WAREHOUSE_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": f"Valid pairs:\n{options_text}{examples_text}\n\nStores:\n{stores_text}",
                    },
                ],
                max_completion_tokens=4000,
                temperature=0.2,
                response_format={"type": "json_object"},
            )
            parsed = json.loads(response.choices[0].message.content or "{}")
            raw_suggestions = parsed.get("suggestions")
            if not isinstance(raw_suggestions, list):
                raise ValueError("Model response didn't match the expected shape")

            for s in raw_suggestions:
                idx = s.get("index")
                if not isinstance(idx, int) or not (1 <= idx <= len(chunk)):
                    continue
                type_code = s.get("warehouse_type_code")
                wh_code = s.get("warehouse_code")
                reasoning = str(s.get("reasoning") or "")
                if (type_code, wh_code) not in valid_set:
                    type_code, wh_code = None, None
                    reasoning = reasoning or "AI suggestion didn't match a configured code -- assign manually"
                result[chunk_start + idx - 1] = (type_code, wh_code, reasoning)
        except Exception:
            logger.exception(
                "Raw import warehouse suggestion request failed for one chunk -- those rows are left blank for manual entry"
            )
            # Other chunks still get their own attempt -- one bad chunk
            # shouldn't blank out the whole batch.
            continue

    return result


def upload_raw_import(
    db: Session, uploaded_by: uuid.UUID, site_id: uuid.UUID, filename: str, content: bytes
) -> tuple[RawImportBatch, list[RawWarehouseSuggestion]]:
    site = db.get(Site, site_id)
    if site is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown site")

    try:
        rows = _parse_raw_rows(content)
    except UploadRejected as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))

    batch = RawImportBatch(site_id=site_id, uploaded_by=uploaded_by, file_name=filename)
    db.add(batch)
    db.flush()

    rows_by_store: dict[tuple[str, str], list[dict]] = {}
    for row in rows:
        rows_by_store.setdefault((row["code_store"], row["store"]), []).append(row)

    raw_candidates = [
        {
            "legacy_code": code_store,
            "legacy_name": store,
            "raw_rows": [
                {"code_rack": r["code_rack"], "description": r["description"], "is_active": r["is_active"]}
                for r in group_rows
            ],
        }
        for (code_store, store), group_rows in rows_by_store.items()
    ]

    # Consolidate billing/status-variant legacy names into one candidate per
    # real physical warehouse before suggesting Type/Code -- see
    # _suggest_clusters and the RawWarehouseSuggestion model docstring.
    cluster_groups = _suggest_clusters(raw_candidates)
    candidates = _build_clustered_candidates(raw_candidates, cluster_groups)

    suggested = _suggest_warehouse_codes(db, candidates)

    suggestions = []
    for candidate, (type_code, wh_code, reasoning) in zip(candidates, suggested):
        suggestion = RawWarehouseSuggestion(
            batch_id=batch.id,
            legacy_code=candidate["legacy_code"],
            legacy_name=candidate["legacy_name"],
            consolidated_legacy_names=candidate["consolidated_legacy_names"],
            raw_rows=candidate["raw_rows"],
            suggested_warehouse_type_code=type_code,
            suggested_warehouse_code=wh_code,
            reasoning=reasoning,
        )
        db.add(suggestion)
        suggestions.append(suggestion)

    db.commit()
    db.refresh(batch)
    for s in suggestions:
        db.refresh(s)
    return batch, suggestions


def _get_pending_warehouse_suggestion(db: Session, suggestion_id: uuid.UUID) -> RawWarehouseSuggestion:
    suggestion = db.get(RawWarehouseSuggestion, suggestion_id)
    if suggestion is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Suggestion not found")
    if suggestion.status != "pending":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Suggestion already resolved")
    return suggestion


def approve_warehouse_suggestion(
    db: Session,
    suggestion_id: uuid.UUID,
    warehouse_type_code: str | None,
    warehouse_code: str | None,
    name: str | None,
    description: str | None,
    capacity: int | None,
) -> RawWarehouseSuggestion:
    suggestion = _get_pending_warehouse_suggestion(db, suggestion_id)
    batch = db.get(RawImportBatch, suggestion.batch_id)

    final_type = warehouse_type_code or suggestion.suggested_warehouse_type_code
    final_code = warehouse_code or suggestion.suggested_warehouse_code
    if not final_type or not final_code:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="warehouse_type_code and warehouse_code are required (the AI had no confident suggestion for this row)",
        )

    warehouse = create_warehouse(
        db, batch.site_id, final_type, final_code, name or suggestion.legacy_name, description, capacity
    )

    suggestion.suggested_warehouse_type_code = final_type
    suggestion.suggested_warehouse_code = final_code
    suggestion.created_warehouse_id = warehouse.id
    suggestion.status = "approved"
    suggestion.resolved_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(suggestion)
    return suggestion


def reject_warehouse_suggestion(db: Session, suggestion_id: uuid.UUID) -> RawWarehouseSuggestion:
    suggestion = _get_pending_warehouse_suggestion(db, suggestion_id)
    suggestion.status = "rejected"
    suggestion.resolved_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(suggestion)
    return suggestion


# --- location-level suggestions ---

_LOCATION_SYSTEM_PROMPT = """You map legacy warehouse rack descriptions to a \
fixed Category Rack scheme. You are given the valid Category Rack values \
for each warehouse type, then a numbered list of racks (each tagged with \
its warehouse type). For each rack, pick the single best-matching valid \
Category Rack value for THAT rack's type -- you may ONLY use a value from \
that type's list, never invent one. If nothing genuinely fits, use null. \
Keep reasoning to a short phrase, under 8 words. You MUST return exactly \
one entry for EVERY numbered rack, in any order, each carrying its own \
"index" (the rack's number) so entries can be matched back correctly even \
if reordered.

You may also be given a list of confirmed examples -- real mappings an \
admin has already reviewed and approved, each tagged with its warehouse \
type. Treat an exact or near-exact description match against a confirmed \
example for the SAME type as very strong evidence, stronger than your own \
prior guess -- but you must still only pick from that type's valid list.

Respond with ONLY a JSON object of this exact shape:
{"suggestions": [{"index": 1, "category_rack": "DRUGS" or null, "reasoning": "..."}]}"""


def _suggest_category_racks(
    db: Session, entries: list[tuple[RawWarehouseSuggestion, Warehouse, dict]]
) -> list[tuple[str | None, str]]:
    """Same chunked, index-matched approach as _suggest_warehouse_codes --
    see that function's docstring and the module docstring for why."""
    default = (None, "AI unavailable -- assign manually")
    result: list[tuple[str | None, str]] = [default for _ in entries]
    if not entries or not settings.groq_api_key:
        return result

    all_mappings = list(db.scalars(select(CategoryRackMapping)).all())
    by_type: dict[str, set[str]] = {}
    for m in all_mappings:
        by_type.setdefault(m.warehouse_type_code, set()).add(m.raw_category_text)
    legend_text = "\n".join(f"Type {t}: {', '.join(sorted(cats))}" for t, cats in sorted(by_type.items()))

    examples = confirmed_location_examples(db)
    examples_text = ""
    if examples:
        examples_text = "\n\nConfirmed examples (already reviewed and approved by an admin):\n" + "\n".join(
            f"- [Type {t}] {desc} -> {cat}" for t, desc, cat in examples
        )

    client = Groq(api_key=settings.groq_api_key)

    for chunk_start, chunk in zip(range(0, len(entries), _SUGGESTION_CHUNK_SIZE), _chunks(entries, _SUGGESTION_CHUNK_SIZE)):
        rows_text = "\n".join(
            f"{i + 1}. [Type {warehouse.warehouse_type_code}] {raw_row['description']}"
            for i, (_ws, warehouse, raw_row) in enumerate(chunk)
        )
        try:
            response = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {"role": "system", "content": _LOCATION_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": f"Valid Category Rack values per warehouse type:\n{legend_text}{examples_text}\n\nRacks:\n{rows_text}",
                    },
                ],
                max_completion_tokens=4000,
                temperature=0.2,
                response_format={"type": "json_object"},
            )
            parsed = json.loads(response.choices[0].message.content or "{}")
            raw_suggestions = parsed.get("suggestions")
            if not isinstance(raw_suggestions, list):
                raise ValueError("Model response didn't match the expected shape")

            for s in raw_suggestions:
                idx = s.get("index")
                if not isinstance(idx, int) or not (1 <= idx <= len(chunk)):
                    continue
                _ws, warehouse, _raw_row = chunk[idx - 1]
                category_rack = s.get("category_rack")
                reasoning = str(s.get("reasoning") or "")
                valid_options = by_type.get(warehouse.warehouse_type_code, set())
                if category_rack not in valid_options:
                    category_rack = None
                result[chunk_start + idx - 1] = (category_rack, reasoning)
        except Exception:
            logger.exception(
                "Raw import location suggestion request failed for one chunk -- those rows are left blank for manual entry"
            )
            continue

    return result


def generate_location_suggestions(db: Session, batch_id: uuid.UUID) -> list[RawLocationSuggestion]:
    batch = db.get(RawImportBatch, batch_id)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Batch not found")

    approved = list(
        db.scalars(
            select(RawWarehouseSuggestion).where(
                RawWarehouseSuggestion.batch_id == batch_id, RawWarehouseSuggestion.status == "approved"
            )
        ).all()
    )
    already_done_ids = set(
        db.scalars(
            select(RawLocationSuggestion.warehouse_suggestion_id).where(RawLocationSuggestion.batch_id == batch_id)
        ).all()
    )
    pending_warehouses = [a for a in approved if a.id not in already_done_ids]
    if not pending_warehouses:
        return []

    entries: list[tuple[RawWarehouseSuggestion, Warehouse, dict]] = []
    for ws in pending_warehouses:
        warehouse = db.get(Warehouse, ws.created_warehouse_id)
        for raw_row in ws.raw_rows:
            entries.append((ws, warehouse, raw_row))

    suggested = _suggest_category_racks(db, entries)

    created = []
    for (ws, warehouse, raw_row), (category_rack, reasoning) in zip(entries, suggested):
        suggestion = RawLocationSuggestion(
            batch_id=batch_id,
            warehouse_suggestion_id=ws.id,
            warehouse_id=warehouse.id,
            legacy_code=raw_row.get("code_rack"),
            legacy_description=raw_row["description"],
            is_active_raw=raw_row.get("is_active", True),
            suggested_category_rack=category_rack,
            reasoning=reasoning,
        )
        db.add(suggestion)
        created.append(suggestion)

    db.commit()
    for s in created:
        db.refresh(s)
    return created


def _get_pending_location_suggestion(db: Session, suggestion_id: uuid.UUID) -> RawLocationSuggestion:
    suggestion = db.get(RawLocationSuggestion, suggestion_id)
    if suggestion is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Suggestion not found")
    if suggestion.status != "pending":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Suggestion already resolved")
    return suggestion


def approve_location_suggestion(
    db: Session, suggestion_id: uuid.UUID, category_rack: str | None, description: str | None
) -> RawLocationSuggestion:
    suggestion = _get_pending_location_suggestion(db, suggestion_id)

    final_category = category_rack or suggestion.suggested_category_rack
    if not final_category:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="category_rack is required (the AI had no confident suggestion for this row)",
        )
    final_description = description or suggestion.legacy_description

    warehouse = db.get(Warehouse, suggestion.warehouse_id)
    mapping = db.scalar(
        select(CategoryRackMapping).where(
            CategoryRackMapping.warehouse_type_code == warehouse.warehouse_type_code,
            CategoryRackMapping.raw_category_text == final_category.upper(),
        )
    )
    if mapping is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"'{final_category}' is not a configured Category Rack mapping for warehouse type "
            f"'{warehouse.warehouse_type_code}'",
        )
    loc_type = db.get(LocationTypeConfig, mapping.location_type_config_id)

    already_exists = db.scalar(
        select(Location).where(
            Location.warehouse_id == warehouse.id,
            Location.location_type_code == loc_type.code,
            Location.description == final_description,
        )
    )
    if already_exists is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT, detail=f"Duplicate: '{final_description}' already exists for this warehouse"
        )

    similar_location, similarity = find_similar_location(db, warehouse.id, loc_type.code, final_description)
    if similar_location is not None:
        merge_suggestion = create_suggestion(
            db, None, None, warehouse.id, loc_type.code, final_category, final_description, similar_location, similarity
        )
        suggestion.created_merge_suggestion_id = merge_suggestion.id
    else:
        location = create_location(db, warehouse.id, loc_type.code, final_category, final_description)
        suggestion.created_location_id = location.id

    suggestion.suggested_category_rack = final_category
    suggestion.status = "approved"
    suggestion.resolved_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(suggestion)
    return suggestion


def reject_location_suggestion(db: Session, suggestion_id: uuid.UUID) -> RawLocationSuggestion:
    suggestion = _get_pending_location_suggestion(db, suggestion_id)
    suggestion.status = "rejected"
    suggestion.resolved_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(suggestion)
    return suggestion
