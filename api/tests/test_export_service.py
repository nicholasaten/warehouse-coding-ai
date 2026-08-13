"""Self-tests the Hospital Unit export against an isolated in-memory
SQLite DB. The core claim under test is round-trip fidelity: whatever
export_site_master produces must be re-ingestible by the exact same
excel_ingest_service functions the Uploads page already uses, with zero
reformatting."""

import io
import uuid

import openpyxl
import pytest
from fastapi import HTTPException
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.db.base import Base
from app.models.category_rack_mapping import CategoryRackMapping
from app.models.location import Location
from app.models.location_type import LocationTypeConfig
from app.models.site import Site
from app.models.warehouse import Warehouse
from app.models.warehouse_code import WarehouseCodeConfig
from app.models.warehouse_type import WarehouseTypeConfig
from app.services.excel_ingest_service import ingest_location_master, ingest_warehouse_master
from app.services.export_service import export_site_master
from app.services.location_service import create_location
from app.services.warehouse_service import create_warehouse

import app.models  # noqa: F401


def _seed_config(db):
    from app.models.user import User

    admin = User(full_name="Test Admin", email="admin@test.com", password_hash="x")
    db.add(admin)
    site = Site(code="RSUS", name="Rumah Sakit Umum Siloam", short_code="US")
    db.add(site)
    db.add(WarehouseTypeConfig(code="A", description="Non-General Items"))
    db.add(WarehouseCodeConfig(warehouse_type_code="A", code="01", description="Pharmacy Mainstore"))
    drugs = LocationTypeConfig(warehouse_type_code="A", code="A", description="Drugs/Consumables")
    db.add(drugs)
    db.flush()
    db.add(CategoryRackMapping(warehouse_type_code="A", raw_category_text="DRUGS", location_type_config_id=drugs.id))
    db.commit()
    return admin, site


@pytest.fixture()
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    session = Session(engine)
    yield session
    session.close()


@pytest.fixture()
def seeded(db):
    return _seed_config(db)


def _read_sheet(content: bytes, sheet_name: str) -> list[tuple]:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb[sheet_name]
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def _single_sheet_bytes(content: bytes, sheet_name: str) -> bytes:
    """excel_ingest_service reads whatever sheet is first in the file
    (pandas.read_excel defaults to sheet_name=0), so re-ingesting one
    sheet out of this export's combined 2-sheet workbook means handing
    it a standalone single-sheet file, not the original bytes."""
    source_ws = openpyxl.load_workbook(io.BytesIO(content))[sheet_name]
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in source_ws.iter_rows(values_only=True):
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_export_site_master_produces_both_sheets(db, seeded):
    _admin, site = seeded
    wh = create_warehouse(db, site.id, "A", "01", "Pharmacy Mainstore Drugs", None, None)
    create_location(db, wh.id, "A", "DRUGS", "DRUGS - TABLET")

    content, filename = export_site_master(db, site.id)

    assert filename == "RSUS-master-export.xlsx"
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb.sheetnames == ["Warehouse Master", "Location Master"]

    wh_rows = _read_sheet(content, "Warehouse Master")
    assert wh_rows[0] == ("Site Code", "Warehouse Type Code", "Warehouse Code", "Warehouse Name", "Description", "Capacity", "Generated Code")
    # openpyxl reads an empty-string cell back as None, not "" -- harmless
    # for re-ingest since excel_ingest_service already treats None as
    # "not provided" the same way it would treat "".
    assert wh_rows[1] == ("RSUS", "A", "01", "Pharmacy Mainstore Drugs", None, None, "RSUS-A01")

    loc_rows = _read_sheet(content, "Location Master")
    assert loc_rows[0] == ("Warehouse Code", "Category Rack", "Description", "Generated Code")
    assert loc_rows[1] == ("RSUS-A01", "DRUGS", "DRUGS - TABLET", "USA01-A01")


def test_export_site_master_empty_site_still_produces_headers(db, seeded):
    _admin, site = seeded
    content, _filename = export_site_master(db, site.id)
    assert _read_sheet(content, "Warehouse Master") == [
        ("Site Code", "Warehouse Type Code", "Warehouse Code", "Warehouse Name", "Description", "Capacity", "Generated Code")
    ]
    assert _read_sheet(content, "Location Master") == [
        ("Warehouse Code", "Category Rack", "Description", "Generated Code")
    ]


def test_export_site_master_not_found(db, seeded):
    with pytest.raises(HTTPException) as exc_info:
        export_site_master(db, uuid.uuid4())
    assert exc_info.value.status_code == 404


def test_export_round_trips_through_the_real_upload_pipeline(db, seeded):
    """The actual point of this feature: export a site, wipe the slate,
    re-ingest the exported file through the same ingest_warehouse_master/
    ingest_location_master the Uploads page uses, and end up with the
    exact same generated codes -- proving the format really is
    importable, not just superficially similar."""
    admin, site = seeded
    wh1 = create_warehouse(db, site.id, "A", "01", "Pharmacy Mainstore Drugs", "orig desc", 15)
    wh2 = create_warehouse(db, site.id, "A", "01", "Pharmacy Mainstore Consumables", None, None)
    create_location(db, wh1.id, "A", "DRUGS", "DRUGS - PSIKOTROPIKA")
    create_location(db, wh2.id, "A", "DRUGS", "DRUGS - TABLET")

    content, _filename = export_site_master(db, site.id)

    # Fresh database, same config, nothing else -- the export must be
    # able to fully recreate the warehouse/location state on its own.
    fresh_engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(fresh_engine)
    fresh_db = Session(fresh_engine)
    try:
        fresh_admin, _fresh_site = _seed_config(fresh_db)

        wh_batch = ingest_warehouse_master(
            fresh_db, fresh_admin.id, "wh.xlsx", _single_sheet_bytes(content, "Warehouse Master")
        )
        assert wh_batch.status == "completed"
        assert wh_batch.success_count == 2
        assert wh_batch.error_count == 0

        loc_batch = ingest_location_master(
            fresh_db, fresh_admin.id, "loc.xlsx", _single_sheet_bytes(content, "Location Master")
        )
        assert loc_batch.status == "completed"
        assert loc_batch.success_count == 2
        assert loc_batch.error_count == 0

        recreated_codes = sorted(w.generated_code for w in fresh_db.scalars(select(Warehouse)).all())
        original_codes = sorted(w.generated_code for w in db.scalars(select(Warehouse)).all())
        assert recreated_codes == original_codes == ["RSUS-A01A", "RSUS-A01B"]

        recreated_loc_codes = sorted(l.generated_code for l in fresh_db.scalars(select(Location)).all())
        original_loc_codes = sorted(l.generated_code for l in db.scalars(select(Location)).all())
        assert recreated_loc_codes == original_loc_codes
    finally:
        fresh_db.close()
