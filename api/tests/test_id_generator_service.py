"""Every literal expected string in this file is copied directly from the
real company data supplied for this project (Formula Warehouse and
Location.pptx, RSUS Mapping.xlsx) -- not invented. This is the ground truth
the whole rule engine is checked against."""

from pathlib import Path

import openpyxl
import pytest

from app.services.id_generator_service import (
    STAGING_RECEIVE_SEQ,
    assign_duplicate_letters,
    format_location_code,
    format_warehouse_code,
    next_location_sequence,
)

SAMPLE_DATA_DIR = Path(__file__).resolve().parents[2] / "sample-data"


# --- Warehouse ID: real examples from the PPTX + RSUS Mapping "WH" sheet ---


@pytest.mark.parametrize(
    "site_code,wh_type,wh_code,dup,expected",
    [
        ("SHBP", "B", "01", "A", "SHBP-B01A"),  # PPTX slide 1 (General Items)
        ("SHMD", "A", "01", "A", "SHMD-A01A"),  # PPTX slide 2 (Non-General Items)
        ("RSUS", "A", "01", "A", "RSUS-A01A"),  # Mapping WH row 2: Pharmacy Mainstore Drugs
        ("RSUS", "A", "01", "B", "RSUS-A01B"),  # Mapping WH row 6: Pharmacy Mainstore Consumables
        ("RSUS", "A", "02", None, "RSUS-A02"),  # Mapping WH row 9: Pharmacy Outpatient (no duplicate)
        ("RSUS", "A", "03", "A", "RSUS-A03A"),  # Mapping WH row 10: Pharmacy Inpatient
        ("RSUS", "A", "03", "B", "RSUS-A03B"),  # Mapping WH row 11: Pharmacy Inpatient Chemotherapy
    ],
)
def test_format_warehouse_code_matches_real_data(site_code, wh_type, wh_code, dup, expected):
    assert format_warehouse_code(site_code, wh_type, wh_code, dup) == expected


# --- Location ID: real examples from RSUS Mapping "LOC" sheet, using the
# "F&O Location Code (Validasi Physical SEQ)" column -- the deduplicated,
# merged final code, which is the actual result the user wants. ---


@pytest.mark.parametrize(
    "site_short,wh_type,wh_code,wh_dup,loc_type,seq,expected",
    [
        ("US", "A", "01", "A", "A", 1, "USA01A-A01"),  # row 2/3 merged: DRUGS - PSIKOTROPIKA
        ("US", "A", "01", "A", "A", 2, "USA01A-A02"),  # row 4/5 merged: DRUGS - OBAT-OBAT TERTENTU
        ("US", "A", "01", "A", "B", 1, "USA01A-B01"),  # row 43/44 merged: COLD STORAGE - KULKAS
        ("US", "A", "01", "A", "E", 1, "USA01A-E01"),  # row 45: EXPIRED ITEMS - RETUR KE SUPPLIER
        ("US", "A", "01", "A", "G", 1, "USA01A-G01"),  # row 47: QUARANTINE - RAK 1
        ("US", "A", "01", "B", "A", 99, "USA01B-A99"),  # rows 65/66: STAGING RECEIVE
        ("US", "A", "02", None, "A", 7, "USA02-A07"),  # row 74: CONSUMABLES - ALKES (no wh duplicate)
        ("US", "A", "04", "A", "D", 1, "USA04A-D01"),  # row 138: BAG - KIT RESUSITASI
        # General Items ("ALL" / whole-warehouse) locations omit the seq entirely:
        ("US", "B", "12", None, "B", None, "USB12-B"),  # row 312: ANCILLARY SERVICES AND MEDICAL AFFAIR - ALL
        ("US", "B", "13", "A", "B", None, "USB13A-B"),  # row 313: NURSING - ALL
        ("US", "B", "18", "A", "B", None, "USB18A-B"),  # rows 322-328: FM & GA A - ALL
    ],
)
def test_format_location_code_matches_real_data(site_short, wh_type, wh_code, wh_dup, loc_type, seq, expected):
    assert format_location_code(site_short, wh_type, wh_code, wh_dup, loc_type, seq) == expected


def test_staging_receive_uses_reserved_seq_99():
    assert STAGING_RECEIVE_SEQ == 99
    assert next_location_sequence(existing_seqs=[1, 2, 3], is_staging_receive=True) == 99


def test_next_location_sequence_increments_and_skips_staging_bucket():
    assert next_location_sequence(existing_seqs=[], is_staging_receive=False) == 1
    assert next_location_sequence(existing_seqs=[1, 2, 3], is_staging_receive=False) == 4
    # 99 (staging) already used shouldn't push the normal counter to 100
    assert next_location_sequence(existing_seqs=[1, 2, 99], is_staging_receive=False) == 3


def test_assign_duplicate_letters_lone_warehouse_gets_no_suffix():
    # Confirmed by "RSUS-A02" in the real data -- only one warehouse under
    # that type+code, so it carries no duplicate letter at all.
    assert assign_duplicate_letters(1) == [None]


def test_assign_duplicate_letters_group_gets_lettered_from_a():
    # Confirmed by "RSUS-A01A" / "RSUS-A01B" both being lettered -- as soon
    # as a second warehouse joins the group, EVERY member gets a letter,
    # not just the new one.
    assert assign_duplicate_letters(2) == ["A", "B"]
    assert assign_duplicate_letters(3) == ["A", "B", "C"]


def test_assign_duplicate_letters_zero_is_empty():
    assert assign_duplicate_letters(0) == []


def test_assign_duplicate_letters_refuses_more_than_26():
    with pytest.raises(ValueError):
        assign_duplicate_letters(27)


# --- Exhaustive regression sweep against the full real reference file, not
# just hand-picked examples -- this is what actually caught the "whole
# warehouse locations omit their seq" edge case during development. Reads
# `sample-data/RSUS Mapping.xlsx` directly rather than hardcoding rows, so
# it stays correct if that reference file is ever updated. ---


def test_every_real_warehouse_id_regenerates_exactly():
    wb = openpyxl.load_workbook(SAMPLE_DATA_DIR / "RSUS Mapping.xlsx", data_only=True)
    ws = wb["WH"]
    checked = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        _hope_code, _name, fo_code, _to_be = row
        if not fo_code or fo_code == "NULL":
            continue
        checked += 1
        site_code, wh_part = fo_code.split("-")
        wh_type, wh_code, dup = wh_part[0], wh_part[1:3], (wh_part[3:] or None)
        assert format_warehouse_code(site_code, wh_type, wh_code, dup) == fo_code, fo_code
    assert checked > 50  # sanity check the file actually has real rows to sweep


def test_every_real_location_id_regenerates_exactly():
    wb = openpyxl.load_workbook(SAMPLE_DATA_DIR / "RSUS Mapping.xlsx", data_only=True)
    ws = wb["LOC"]
    checked = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        validasi_code = row[10]  # "F&O Location Code (Validasi Physical SEQ)" -- the deduplicated final code
        if not validasi_code:
            continue
        checked += 1
        wh_part, loc_part = validasi_code.split("-")
        loc_type_code, seq_str = loc_part[0], loc_part[1:]
        seq = int(seq_str) if seq_str else None
        site_short, wh_type, wh_code = wh_part[:2], wh_part[2], wh_part[3:5]
        dup = wh_part[5:] or None
        regenerated = format_location_code(site_short, wh_type, wh_code, dup, loc_type_code, seq)
        assert regenerated == validasi_code, validasi_code
    assert checked > 300  # sanity check the file actually has real rows to sweep
